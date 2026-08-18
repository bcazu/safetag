"""Genera kobo/ais.xlsx — Formulario Único AIS (v2) para KoboToolbox.

Digitaliza el *Formulario Único de Inspección de Edificaciones Después de un
Sismo* (AIS, Comité AIS-400). Fuente primaria: manual completo en
docs/referencias/manual-ais-manizales-2003.pdf; tablas verificadas en
docs/manual-ais-tablas-verificadas.md.
Secciones cubiertas: 1 (catastral), 2 (tipo inspección), 3 (identificación),
4 (estructura), 5.1 (estabilidad), 5.2 (geotécnica observable), 5.3 (daños
estructurales), 6 (% global), 11 (ocupación no sensible), 13 (comentarios),
14 (identificación del brigadista), 17 (fotos). Pendiente T6 (consentimiento
expreso) para sumar 10 y 12 — datos sensibles/PII del ocupante.

El brigadista DOCUMENTA, no evalúa: el formulario captura observaciones; los
riesgos y la habitabilidad los asigna el revisor profesional (packages/rules).
Las guías de niveles de daño provienen textualmente de las tablas 3-4 a 3-11
del manual — son ayudas de observación, no criterios nuevos.

Uso:
    pip install openpyxl pyxform   # en un venv
    python kobo/gen_xlsform.py
    xls2xform kobo/ais.xlsx /tmp/ais.xml   # validar

Listas territoriales: kobo/media/*.csv (select_one_from_file). Se suben a
Kobo como media files del formulario (Settings → Media); corregirlas no
exige redesplegar. Ver kobo/media/README.md (fuentes y TODOs).

Multiidioma: columnas `label::Idioma (cod)`. Para sumar un idioma, añadirlo a
LANGS y completar los diccionarios {"es": ...}. Los `name` de los campos son
el contrato con supabase/functions/kobo-webhook: no renombrar a la ligera.
El formulario agrupa en páginas (settings.style = "pages"): Kobo entrega los
campos como `grupo/campo`, el webhook resuelve con su helper field().
"""
from pathlib import Path

from openpyxl import Workbook

LANGS = {"es": "Español (es)"}
DEFAULT_LANG = "Español (es)"

I18N_COLS = ["label", "hint", "constraint_message"]
PLAIN_COLS = ["type", "name", "required", "appearance", "relevant",
              "constraint", "parameters", "choice_filter", "default",
              "calculation"]

# Default de municipio por despliegue (código DIVIPOLA) — parámetro, no
# hardcode: el form desplegado para la brigada de Pereira puede llevar "66001".
MUNICIPIO_DEFAULT = ""

# Relevancia común: todo salvo intro/inspección/ubicación/fotos se oculta si
# no se inspeccionó (se aplica a nivel de grupo/página).
INSPECTED = "${tipo_inspeccion} != 'no_inspeccionada'"

# Conjuntos de sistema estructural para la matriz de daño (sección 5.3)
PORTICOS = ["11", "12", "13", "14", "31", "32", "33", "41", "42"]
NUDOS = ["11", "12", "13", "14"]
CONEXIONES = ["31", "32", "33", "41", "42"]
MUROS_CR = ["12", "13"]
MUROS_PORTANTES = ["21", "22", "23", "51", "52", "50", "60"]
CONFINADA = ["21"]


def sys_in(codes):
    ors = " or ".join(f"${{sistema_estructural}} = '{c}'" for c in codes)
    return f"({ors})"


def damage_pair(name, label_es, codes):
    """select_one nivel + % de extensión para un elemento de la matriz 5.3.

    La relevancia por INSPECTED y colapso != total la aporta el grupo `danos`.
    """
    rel = sys_in(codes)
    return [
        {"type": "select_one niveles_dano", "name": f"dano_{name}",
         "label": {"es": f"Daño en {label_es}"}, "required": "yes",
         "relevant": rel},
        {"type": "integer", "name": f"dano_{name}_pct",
         "label": {"es": f"% de extensión del daño en {label_es}"},
         "required": "yes",
         "relevant": f"{rel} and ${{dano_{name}}} != 'ninguno'",
         "constraint": ". >= 1 and . <= 100",
         "constraint_message": {"es": "Debe estar entre 1 y 100"},
         "hint": {"es": f"De cada 10 {label_es} de ese piso, ¿cuántos tienen "
                        "ese daño? 3 de cada 10 = 30"}},
    ]


def guia_dano(name, codes, texto_es):
    """Nota de guía de niveles de daño por material (tablas 3-4 a 3-11 AIS)."""
    return {"type": "note", "name": f"guia_{name}",
            "label": {"es": texto_es}, "relevant": sys_in(codes)}


survey = [
    # metadatos
    {"type": "start", "name": "start"},
    {"type": "end", "name": "end"},
    {"type": "start-geopoint", "name": "gps_fondo"},

    # ── Página 1: antes de empezar ───────────────────────────────────────
    {"type": "begin_group", "name": "intro",
     "label": {"es": "Antes de empezar"}, "appearance": "field-list"},
    {"type": "note", "name": "nota_rol",
     "label": {"es": "USTED DOCUMENTA, NO EVALÚA. Registre lo que observa y "
                     "fotografíe: un ingeniero matriculado revisará este "
                     "reporte y asignará la clasificación. Si tiene dudas en "
                     "una pregunta, elija la opción más cercana a lo que ve "
                     "y anótelo en «Comentarios» al final."}},
    {"type": "note", "name": "nota_seguridad",
     "label": {"es": "⚠️ SU SEGURIDAD PRIMERO. No entre a edificaciones con "
                     "daño evidente ni permanezca bajo balcones, fachadas o "
                     "elementos que puedan caer. Todo lo importante se puede "
                     "documentar desde el exterior."}},
    # aviso de privacidad + confirmación (T6; política completa en
    # docs/politica-datos.md). Cuando se agreguen las secciones 10 y 12
    # (ocupantes/contacto), añadir aquí el consentimiento expreso para datos
    # sensibles, con mención de que el titular NO está obligado a autorizarlo.
    {"type": "note", "name": "aviso_privacidad",
     "label": {"es": "AVISO DE PRIVACIDAD — Esta inspección registra "
                     "ubicación, características y fotografías de la "
                     "edificación (solo exteriores y elementos "
                     "estructurales, no personas ni interiores habitados "
                     "salvo lo estrictamente necesario) para priorizar la "
                     "evaluación estructural post-sismo. Tratamiento según "
                     "la Ley 1581 de 2012; política completa disponible con "
                     "la entidad coordinadora. La clasificación resultante "
                     "es un insumo de priorización: la habilitación "
                     "definitiva es competencia de las autoridades."}},
    {"type": "select_one confirmaciones", "name": "aviso_comunicado",
     "label": {"es": "¿Informaste al ocupante o responsable del inmueble el "
                     "propósito de la inspección y el aviso de privacidad?"},
     "required": "yes",
     "hint": {"es": "Si no hay nadie presente, marca «No aplica (inmueble "
                    "desocupado o sin acceso)»"}},
    # Sección 14 AIS (identificación del inspector) adaptada a brigadas:
    # trazabilidad de quién documenta, sin PII de terceros.
    {"type": "text", "name": "codigo_comision",
     "label": {"es": "Código de tu comisión o brigada"},
     "hint": {"es": "El asignado por la coordinación (ej.: BRG-05). Déjalo "
                    "vacío si no te asignaron uno."}},
    {"type": "text", "name": "codigo_brigadista",
     "label": {"es": "Tu código de brigadista o nombre completo"},
     "required": "yes",
     "hint": {"es": "Queda como responsable de la documentación de esta "
                    "inspección"}},
    {"type": "end_group", "name": ""},

    # ── Página 2: tipo de inspección (sección 2 AIS) ─────────────────────
    {"type": "begin_group", "name": "inspeccion",
     "label": {"es": "Tipo de inspección"}, "appearance": "field-list"},
    {"type": "select_one tipos_inspeccion", "name": "tipo_inspeccion",
     "label": {"es": "¿Qué tipo de inspección pudiste hacer?"},
     "required": "yes",
     "hint": {"es": "«Parcial» = exterior más algunas zonas interiores. Si "
                    "no fue posible inspeccionar, igual registra ubicación "
                    "y fotos: ese dato también sirve."}},
    {"type": "select_one motivos_no_inspeccion", "name": "motivo_no_inspeccion",
     "label": {"es": "¿Por qué no se inspeccionó?"}, "required": "yes",
     "relevant": "${tipo_inspeccion} = 'no_inspeccionada'"},
    {"type": "end_group", "name": ""},

    # ── Página 3: ubicación e identificación (secciones 1 y 3) ──────────
    # El GPS es la fuente de verdad; los nombres son metadatos legibles
    # (reasignación por ST_Contains en gabinete). Listas territoriales como
    # CSV adjuntos (kobo/media/*.csv), NUNCA en la hoja choices: corregir
    # una lista no puede exigir redesplegar el formulario. Los `name` de
    # municipios.csv SON los códigos DIVIPOLA (contrato con la BD).
    {"type": "begin_group", "name": "ubicacion_grupo",
     "label": {"es": "¿Dónde está la edificación?"},
     "appearance": "field-list"},
    {"type": "select_one_from_file municipios.csv", "name": "municipio",
     "label": {"es": "Municipio"}, "required": "yes", "appearance": "minimal",
     "default": MUNICIPIO_DEFAULT},
    {"type": "select_one tipo_zona", "name": "tipo_zona",
     "label": {"es": "Tipo de zona"}, "required": "yes"},
    # Urbano y rural comparten estructura (misma columna en BD): dos
    # preguntas con relevant excluyente en vez de label dinámico — más
    # simple en XLSForm. Comuna ↔ corregimiento.
    {"type": "select_one_from_file divisiones.csv", "name": "division_urbana",
     "label": {"es": "Comuna"}, "required": "yes", "appearance": "minimal",
     "relevant": "${tipo_zona} = 'urbano'",
     "choice_filter": "municipio_code=${municipio} and tipo='urbano'"},
    {"type": "select_one_from_file divisiones.csv", "name": "division_rural",
     "label": {"es": "Corregimiento"}, "required": "yes",
     "appearance": "minimal",
     "relevant": "${tipo_zona} = 'rural'",
     "choice_filter": "municipio_code=${municipio} and tipo='rural'"},
    # Enketo no evalúa if() dentro de un choice_filter (FormLogicError
    # "Too many tokens"): la división efectiva se resuelve aparte y el
    # filtro queda con la gramática plana que sí soporta.
    {"type": "calculate", "name": "division_sel",
     "calculation": "if(${tipo_zona} = 'urbano', "
                    "${division_urbana}, ${division_rural})"},
    # La fila OTRO de barrios.csv es global: el choice_filter la incluye
    # siempre para que ninguna lista incompleta bloquee una evaluación.
    {"type": "select_one_from_file barrios.csv", "name": "barrio",
     "label": {"es": "Barrio / Vereda"}, "required": "yes",
     "appearance": "minimal",
     "choice_filter": "division_code=${division_sel} or name='OTRO'"},
    {"type": "text", "name": "barrio_otro", "label": {"es": "¿Cuál?"},
     "required": "yes", "relevant": "${barrio} = 'OTRO'"},
    {"type": "select_one via_tipo", "name": "via_tipo",
     "label": {"es": "Tipo de vía"},
     "relevant": "${tipo_zona} = 'urbano'"},
    {"type": "text", "name": "via_numero", "label": {"es": "Número de vía"},
     "relevant": "${tipo_zona} = 'urbano'",
     "hint": {"es": "Ej.: para Carrera 7 # 45-12, aquí va el 7"}},
    {"type": "text", "name": "numero_placa", "label": {"es": "Número (placa)"},
     "relevant": "${tipo_zona} = 'urbano'",
     "hint": {"es": "El resto de la dirección (45-12). Incluye interior, "
                    "bloque o torre si es un conjunto."}},
    {"type": "text", "name": "referencia_ubicacion",
     "label": {"es": "Referencia de ubicación"},
     "hint": {"es": "Finca, kilómetro, punto de referencia. Si el sector "
                    "tiene doble nomenclatura, anota la otra dirección en "
                    "«Comentarios»."}},
    {"type": "text", "name": "nombre_edificacion",
     "label": {"es": "Nombre de la edificación (si tiene)"},
     "hint": {"es": "Propiedad horizontal o institución (ej.: Edificio Los "
                    "Cámbulos, I.E. La Julita)"}},
    # Sin constraint que impida guardar por ubicación: el GPS puede fallar
    # bajo techo; el webhook acepta el start-geopoint de respaldo como
    # último recurso. Nunca perder una evaluación por GPS.
    {"type": "geopoint", "name": "ubicacion", "label": {"es": "Ubicación GPS"},
     "required": "yes",
     "hint": {"es": "Capturar de pie frente a la edificación, con cielo "
                    "visible. Si no captura bajo techo, reintentar afuera."}},
    # La identificación catastral (IGAC) NO se le exige al brigadista: el
    # manual AIS asigna el número predial en oficina cruzando con catastro.
    {"type": "begin_group", "name": "catastral",
     "label": {"es": "Identificación catastral (opcional)"},
     "hint": {"es": "Solo si tiene el recibo de predial a la vista"}},
    {"type": "text", "name": "cat_sector", "label": {"es": "Sector"}},
    {"type": "text", "name": "cat_manzana", "label": {"es": "Manzana"}},
    {"type": "text", "name": "cat_predio", "label": {"es": "Predio"}},
    {"type": "text", "name": "cat_mejora",
     "label": {"es": "Mejora / Prop. horizontal"}},
    {"type": "end_group", "name": ""},
    {"type": "end_group", "name": ""},

    # ── Página 4: la edificación (sección 3 continuación + 11) ──────────
    {"type": "begin_group", "name": "edificacion",
     "label": {"es": "La edificación"}, "appearance": "field-list",
     "relevant": INSPECTED},
    {"type": "integer", "name": "pisos_sobre",
     "label": {"es": "Niveles sobre el terreno"}, "required": "yes",
     "constraint": ". >= 1 and . <= 50",
     "constraint_message": {"es": "Debe estar entre 1 y 50"},
     "hint": {"es": "Sin contar cubierta ni terraza. Si está en ladera, "
                    "cuenta desde la entrada principal."}},
    {"type": "integer", "name": "sotanos", "label": {"es": "Sótanos"},
     "constraint": ". >= 0 and . <= 10",
     "constraint_message": {"es": "Debe estar entre 0 y 10"}},
    {"type": "select_one usos", "name": "uso_predominante",
     "label": {"es": "Uso predominante de la edificación"}, "required": "yes",
     "appearance": "minimal",
     "hint": {"es": "Si hay varios usos, el que ocupa la mayor parte"}},
    {"type": "select_one usos", "name": "uso_planta_baja",
     "label": {"es": "Uso de la planta baja"},
     "appearance": "minimal",
     "hint": {"es": "Solo si es distinto del uso predominante (ej.: local "
                    "comercial bajo apartamentos)"}},
    {"type": "decimal", "name": "frente_m",
     "label": {"es": "Frente aproximado (m)"},
     "constraint": ". > 0 and . < 1000",
     "constraint_message": {"es": "Valor en metros, mayor que 0"},
     "hint": {"es": "Si no puede medir: un paso largo ≈ 1 m"}},
    {"type": "decimal", "name": "fondo_m",
     "label": {"es": "Fondo aproximado (m)"},
     "constraint": ". > 0 and . < 1000",
     "constraint_message": {"es": "Valor en metros, mayor que 0"}},
    # Sección 11 (solo el dato no sensible; heridos/fallecidos y contacto
    # esperan el consentimiento expreso de T6)
    {"type": "select_one habitada_opts", "name": "habitada",
     "label": {"es": "¿La edificación está habitada u ocupada actualmente?"},
     "required": "yes"},
    {"type": "end_group", "name": ""},

    # ── Página 5: estructura (sección 4) ─────────────────────────────────
    {"type": "begin_group", "name": "estructura",
     "label": {"es": "¿De qué está hecha?"}, "appearance": "field-list",
     "relevant": INSPECTED},
    {"type": "select_one sistemas_estructurales", "name": "sistema_estructural",
     "label": {"es": "Sistema estructural"}, "required": "yes",
     "appearance": "minimal",
     "hint": {"es": "Guía rápida: ¿esqueleto de columnas y vigas de concreto "
                    "con muros solo de relleno? → Pórtico de concreto. "
                    "¿Muros de ladrillo que cargan la casa, con columnetas y "
                    "vigas de amarre en concreto? → Mampostería confinada; "
                    "sin amarres → no reforzada. ¿Esqueleto metálico? → "
                    "acero. ¿Guadua o esterilla con revoque? → bahareque. "
                    "¿Muros gruesos de tierra pisada? → tapia. Si no logra "
                    "identificarlo, elija «Otros / no identificable» y "
                    "descríbalo en Comentarios."}},
    {"type": "select_one tipos_entrepiso", "name": "tipo_entrepiso",
     "label": {"es": "Tipo de entrepiso"},
     "appearance": "minimal",
     "hint": {"es": "Lo que separa un piso de otro; suele verse desde abajo "
                    "(el «techo» del primer piso). Si no lo distingue, "
                    "«Otros / no se sabe»."}},
    {"type": "select_one rangos_ano", "name": "rango_ano",
     "label": {"es": "Año de construcción"}, "required": "yes",
     "hint": {"es": "Pregunte al ocupante o vecino. Si nadie lo sabe, "
                    "marque «No se sabe»: no adivine."}},
    {"type": "end_group", "name": ""},

    # ── Página 6: estado global y terreno (secciones 5.1 y 5.2) ─────────
    {"type": "begin_group", "name": "estado",
     "label": {"es": "Estado general y terreno"}, "appearance": "field-list",
     "relevant": INSPECTED},
    {"type": "select_one colapsos", "name": "colapso",
     "label": {"es": "Colapso"}, "required": "yes",
     "hint": {"es": "«Parcial» = una parte de la edificación cayó (un piso, "
                    "un ala, la cubierta)"}},
    {"type": "select_one inclinaciones", "name": "inclinacion",
     "label": {"es": "Inclinación de la edificación o de un piso"},
     "required": "yes",
     "hint": {"es": "Compárela con las edificaciones vecinas o con un hilo "
                    "con peso (plomada). Si no está seguro, «Hay dudas»."}},
    # Sección 5.2 — lo observable por el brigadista (asentamiento, talud,
    # morfología). Origen y potencial de reactivación son juicio del
    # profesional y quedan del lado del revisor. Estas respuestas activan el
    # enrutamiento a geotecnista (Ley 842 art. 19, routeCase).
    {"type": "select_one asentamientos", "name": "asentamiento",
     "label": {"es": "Asentamiento o hundimiento de la edificación"},
     "required": "yes",
     "hint": {"es": "¿Se hundió o quedó más baja que el andén o los "
                    "vecinos? ¿Aparecieron desniveles o escalones nuevos "
                    "en el piso?"}},
    {"type": "select_one fallas_talud", "name": "falla_talud",
     "label": {"es": "Falla en talud o movimiento en masa cerca"},
     "required": "yes",
     "hint": {"es": "Deslizamientos, caída de tierra o piedras, o grietas "
                    "en el terreno alrededor de la edificación. «General» = "
                    "afecta toda la ladera o varias edificaciones."}},
    {"type": "select_one morfologias", "name": "morfologia",
     "label": {"es": "¿Cómo es el terreno donde está la edificación?"},
     "appearance": "minimal"},
    {"type": "end_group", "name": ""},

    # ── Página 7: daños estructurales (sección 5.3) ─────────────────────
    {"type": "begin_group", "name": "danos",
     "label": {"es": "Daños en la estructura"}, "appearance": "field-list",
     "relevant": f"{INSPECTED} and ${{colapso}} != 'total'"},
    {"type": "integer", "name": "piso_mayor_dano",
     "label": {"es": "¿Cuál es el piso de mayor daño?"}, "required": "yes",
     "constraint": ". >= 0 and . <= ${pisos_sobre}",
     "constraint_message": {"es": "0 = sótano; no puede ser mayor que el "
                                  "número de niveles"},
     "hint": {"es": "Las siguientes preguntas se responden mirando ese piso"}},
    # Guías de observación por material — texto de las tablas 3-4 a 3-10 del
    # manual AIS (docs/manual-ais-tablas-verificadas.md). No inventar valores.
    guia_dano("concreto", ["11", "12", "13", "14"],
              "GUÍA para concreto — Ninguno: fisuras casi invisibles "
              "(< 0,2 mm). Leve: se ven a simple vista (hasta 1 mm). "
              "Moderado: 1–2 mm o recubrimiento empezando a desprenderse. "
              "Fuerte: recubrimiento caído o varillas a la vista. Severo: "
              "concreto triturado, varillas dobladas, deformación evidente."),
    guia_dano("mamposteria", ["21", "22", "23"],
              "GUÍA para mampostería (ladrillo/bloque) — Ninguno: fisuras "
              "casi invisibles (< 0,2 mm). Leve: hasta 1 mm. Moderado: "
              "1–3 mm o inicio de grietas diagonales. Fuerte: grietas "
              "diagonales > 3 mm o ladrillos dislocados. Severo: piezas "
              "desprendidas, aplastamiento o muro desplomado."),
    guia_dano("acero", ["31", "32", "33"],
              "GUÍA para acero — Leve: deformaciones apenas perceptibles. "
              "Moderado: deformación visible o pandeo incipiente. Fuerte/"
              "Severo: pandeo o fractura clara, o soldaduras, tornillos o "
              "remaches rotos. Revise primero las conexiones y fotografíelas."),
    guia_dano("madera", ["41", "42"],
              "GUÍA para madera — Leve: fisuras mínimas. Moderado: grietas "
              "con uniones aún firmes. Fuerte: uniones claramente corridas "
              "o desplazadas. Severo: pieza rota, seccionada o suelta de la "
              "estructura."),
    guia_dano("bahareque", ["51"],
              "GUÍA para bahareque — Leve: grietas incipientes del revoque. "
              "Moderado: grietas en esquinas y revoque desprendiéndose. "
              "Fuerte: grietas en casi todos los muros o cubierta perdiendo "
              "apoyo. Severo: muro deformado o pandeado, pies derechos "
              "sueltos."),
    guia_dano("tapia", ["52"],
              "GUÍA para tapia/adobe — Ninguno: fisuras < 0,4 mm. Leve: "
              "0,4–2 mm. Moderado: 2–4 mm. Fuerte: > 4 mm o muro movido "
              "fuera de su plano. Severo: aplastamiento o desplome."),
    *damage_pair("vigas", "vigas", PORTICOS),
    *damage_pair("columnas", "columnas", PORTICOS),
    *damage_pair("nudos", "nudos (unión viga-columna)", NUDOS),
    *damage_pair("conexiones", "conexiones", CONEXIONES),
    *damage_pair("muros", "muros estructurales", MUROS_CR),
    *damage_pair("muros_portantes", "muros portantes", MUROS_PORTANTES),
    *damage_pair("columnetas", "columnetas de confinamiento", CONFINADA),
    *damage_pair("vigas_confinamiento", "vigas de confinamiento", CONFINADA),
    *damage_pair("entrepisos", "entrepisos",
                 PORTICOS + MUROS_PORTANTES),
    # Medición objetiva opcional: complementa la foto con escala; el revisor
    # la contrasta con los umbrales por material (packages/rules).
    {"type": "decimal", "name": "ancho_grieta_mm",
     "label": {"es": "Ancho de la peor grieta (mm), si pudo medirlo"},
     "constraint": ". >= 0 and . <= 500",
     "constraint_message": {"es": "Valor en milímetros"},
     "hint": {"es": "Solo si tiene regla o fisurómetro. Si no, déjelo vacío "
                    "y asegure la foto con un objeto de referencia."}},
    # señales rápidas para priorización/IA (no es el registro de daño oficial)
    {"type": "select_multiple senales", "name": "senales_alarma",
     "label": {"es": "Señales de alarma observadas"}, "required": "yes",
     "hint": {"es": "Marcar todas las que apliquen; si no hay, marcar «Ninguna»"},
     "constraint": "not(selected(., 'ninguna') and count-selected(.) > 1)",
     "constraint_message": {"es": "«Ninguna» no puede combinarse con otras señales"}},
    {"type": "end_group", "name": ""},

    # ── Página 8: daño global, peligros exteriores y comentarios ────────
    {"type": "begin_group", "name": "cierre",
     "label": {"es": "Balance general"}, "appearance": "field-list",
     "relevant": INSPECTED},
    {"type": "select_one danos_globales", "name": "dano_global",
     "label": {"es": "Porcentaje global de daño de la edificación"},
     "required": "yes",
     "constraint": "${colapso} != 'total' or . = '100'",
     "constraint_message": {"es": "Si el colapso es total, el daño global "
                                  "es 100 %"},
     "hint": {"es": "Área afectada frente al área total, contando daños "
                    "estructurales y no estructurales. No cuente muebles "
                    "ni enseres."}},
    # Elementos en peligro de caer / redes rotas (indicadores de riesgo no
    # estructural ALTO según tabla 3-21; también insumo de seguridad para
    # el PMU: afectan la vía y a los vecinos).
    {"type": "select_multiple peligros_exterior", "name": "peligros_exterior",
     "label": {"es": "Peligros sobre la calle, vecinos u ocupantes"},
     "required": "yes",
     "hint": {"es": "Marcar todos los que apliquen; si no hay, «Ninguno»"},
     "constraint": "not(selected(., 'ninguno') and count-selected(.) > 1)",
     "constraint_message": {"es": "«Ninguno» no puede combinarse con otros"}},
    # Sección 13
    {"type": "text", "name": "comentarios",
     "label": {"es": "Comentarios"}, "appearance": "multiline",
     "hint": {"es": "Todo lo que no cupo en el formulario: combinación de "
                    "sistemas constructivos, doble nomenclatura, daños que "
                    "no encajan en las preguntas, acceso, etc."}},
    {"type": "end_group", "name": ""},

    # ── Página 9: fotografías (sección 17) ──────────────────────────────
    {"type": "begin_group", "name": "fotos",
     "label": {"es": "Fotografías"}, "appearance": "field-list"},
    {"type": "note", "name": "nota_fotos",
     "label": {"es": "Las fotos son lo más importante del reporte: el "
                     "ingeniero evalúa con ellas. Tómelas con buena luz y "
                     "sin personas en el encuadre."}},
    {"type": "image", "name": "foto_fachada",
     "label": {"es": "Foto: fachada completa"}, "required": "yes",
     "hint": {"es": "Desde el frente, que se vea todo el edificio"},
     "parameters": "max-pixels=1600"},
    {"type": "image", "name": "foto_esquina_1",
     "label": {"es": "Foto: esquina 1"}, "required": "yes",
     "hint": {"es": "Desde una esquina, que se vean dos fachadas a la vez"},
     "parameters": "max-pixels=1600"},
    {"type": "image", "name": "foto_esquina_2",
     "label": {"es": "Foto: esquina 2"}, "required": "yes",
     "hint": {"es": "La esquina opuesta, si puede dar la vuelta"},
     "parameters": "max-pixels=1600"},
    {"type": "image", "name": "foto_esquina_3",
     "label": {"es": "Foto: esquina 3 (si es accesible)"},
     "parameters": "max-pixels=1600"},
    {"type": "image", "name": "foto_esquina_4",
     "label": {"es": "Foto: esquina 4 (si es accesible)"},
     "parameters": "max-pixels=1600"},
    {"type": "image", "name": "foto_grieta",
     "label": {"es": "Foto: peor grieta, de cerca"},
     "required": "yes",
     "relevant": f"{INSPECTED} and ${{dano_global}} != 'ninguno'",
     "hint": {"es": "Poner una moneda o una regla junto a la grieta como "
                    "referencia de escala"},
     "parameters": "max-pixels=2048"},
    {"type": "image", "name": "foto_columnas",
     "label": {"es": "Foto: columnas o muros del primer piso"},
     "required": "yes", "relevant": INSPECTED,
     "hint": {"es": "El primer piso es el que más carga: fotografíe sus "
                    "columnas o muros aunque se vean sanos"},
     "parameters": "max-pixels=1600"},
    {"type": "end_group", "name": ""},
]

# ── Listas de opciones ───────────────────────────────────────────────────────

confirmaciones = [
    ("si", {"es": "Sí"}),
    ("no_aplica", {"es": "No aplica (inmueble desocupado o sin acceso)"}),
]

tipos_inspeccion = [
    ("exterior", {"es": "Solo exterior"}),
    ("parcial", {"es": "Parcial (exterior + algunas zonas interiores)"}),
    ("completa", {"es": "Completa"}),
    ("no_inspeccionada", {"es": "No se inspeccionó"}),
]

motivos_no_inspeccion = [
    ("no_permitido", {"es": "No se permitió el acceso"}),
    ("desocupada", {"es": "Edificación desocupada"}),
    ("colapso", {"es": "Colapso total"}),
    ("demolida", {"es": "Ya demolida"}),
    ("otro", {"es": "Otro"}),
]

# Las listas territoriales (municipios, divisiones, barrios) viven en
# kobo/media/*.csv como media files del formulario (HANDOFF-T5b): corregir
# una lista no exige redesplegar. Aquí solo quedan listas estables.
tipo_zona = [
    ("urbano", {"es": "Urbana"}),
    ("rural", {"es": "Rural"}),
]

via_tipo = [
    ("carrera", {"es": "Carrera"}),
    ("calle", {"es": "Calle"}),
    ("transversal", {"es": "Transversal"}),
    ("diagonal", {"es": "Diagonal"}),
    ("avenida", {"es": "Avenida"}),
    ("otra", {"es": "Otra"}),
]

# Códigos oficiales AIS (sección 4) — el `name` ES el código
sistemas_estructurales = [
    ("11", {"es": "Concreto: pórtico (esqueleto de columnas y vigas)"}),
    ("12", {"es": "Concreto: muros estructurales (muros de concreto cargan)"}),
    ("13", {"es": "Concreto: sistema dual (pórtico + muros)"}),
    ("14", {"es": "Concreto: prefabricado"}),
    ("21", {"es": "Mampostería confinada (ladrillo con vigas y columnas de amarre)"}),
    ("22", {"es": "Mampostería reforzada (bloques con refuerzo interno)"}),
    ("23", {"es": "Mampostería no reforzada (ladrillo sin amarres)"}),
    ("31", {"es": "Acero: pórticos arriostrados (con diagonales)"}),
    ("32", {"es": "Acero: pórticos no arriostrados"}),
    ("33", {"es": "Acero: pórticos en celosía"}),
    ("41", {"es": "Madera: pórticos y panel en madera"}),
    ("42", {"es": "Madera: pórticos con paneles de otros materiales"}),
    ("51", {"es": "Muros en bahareque (guadua/esterilla con revoque)"}),
    ("52", {"es": "Muros en tapia (tierra pisada) o adobe"}),
    ("50", {"es": "Mixta (combinación de sistemas)"}),
    ("60", {"es": "Otros / no identificable"}),
]

tipos_entrepiso = [
    ("11", {"es": "Concreto: placa maciza"}),
    ("12", {"es": "Concreto: placa aligerada"}),
    ("13", {"es": "Concreto: reticular celulado"}),
    ("21", {"es": "Acero: vigas alma llena con conectores"}),
    ("22", {"es": "Acero: vigas alma llena sin conectores"}),
    ("23", {"es": "Acero: cerchas"}),
    ("31", {"es": "Madera: vigas"}),
    ("32", {"es": "Madera: cerchas"}),
    ("40", {"es": "Mixta"}),
    ("50", {"es": "Otros / no se sabe"}),
]

usos = [
    ("1", {"es": "Residencial"}), ("2", {"es": "Comercial"}),
    ("3", {"es": "Educacional"}), ("4", {"es": "Salud"}),
    ("5", {"es": "Hotelero"}), ("6", {"es": "Oficinas"}),
    ("7", {"es": "Industrial"}), ("8", {"es": "Institucional"}),
    ("9", {"es": "Bodegas"}), ("10", {"es": "Parqueaderos"}),
    ("11", {"es": "Otros"}),
]

# Cortes según la entrada en vigor de los códigos sismorresistentes.
# "0 No se sabe" no es un código AIS: es el escape del brigadista lego
# (documenta, no adivina); el webhook lo traduce a NULL.
rangos_ano = [
    ("1", {"es": "Antes de 1950"}),
    ("2", {"es": "1950 – 1982"}),
    ("3", {"es": "1982 – 1997"}),
    ("4", {"es": "1998 en adelante"}),
    ("0", {"es": "No se sabe"}),
]

colapsos = [
    ("total", {"es": "Colapso total"}),
    ("parcial_mayor_50", {"es": "Colapso parcial (50% o más)"}),
    ("parcial_menor_50", {"es": "Colapso parcial (menos del 50%)"}),
    ("ninguno", {"es": "Ninguno"}),
]

inclinaciones = [
    ("evidente", {"es": "Inclinación evidente"}),
    ("dudas", {"es": "Hay dudas"}),
    ("ninguna", {"es": "Ninguna"}),
]

# Sección 5.2 AIS — opciones oficiales del formulario impreso
asentamientos = [
    ("evidente", {"es": "Evidente"}),
    ("dudas", {"es": "Existen dudas"}),
    ("ninguno", {"es": "Ninguno"}),
]

fallas_talud = [
    ("general", {"es": "General (toda la ladera o varias edificaciones)"}),
    ("puntual", {"es": "Puntual"}),
    ("ninguna", {"es": "Ninguna"}),
]

morfologias = [
    ("1", {"es": "Divisoria"}),
    ("2", {"es": "Cresta"}),
    ("3", {"es": "Ladera"}),
    ("4", {"es": "Pie de ladera"}),
    ("5", {"es": "Valle"}),
    ("6", {"es": "Canal"}),
    ("7", {"es": "Borde de río"}),
    ("8", {"es": "Talud"}),
]

habitada_opts = [
    ("si", {"es": "Sí"}),
    ("no", {"es": "No"}),
    ("no_se_sabe", {"es": "No se sabe"}),
]

niveles_dano = [
    ("ninguno", {"es": "Ninguno / muy leve"}),
    ("leve", {"es": "Leve"}),
    ("moderado", {"es": "Moderado"}),
    ("fuerte", {"es": "Fuerte"}),
    ("severo", {"es": "Severo"}),
]

danos_globales = [
    ("ninguno", {"es": "Ninguno"}),
    ("0_10", {"es": "0 – 10 %"}),
    ("10_30", {"es": "10 – 30 %"}),
    ("30_60", {"es": "30 – 60 %"}),
    ("60_100", {"es": "60 – 100 %"}),
    ("100", {"es": "100 % (colapso)"}),
]

senales = [
    ("grietas_x", {"es": "Grietas en X en muros"}),
    ("grietas_horizontales_base", {"es": "Grietas horizontales en base de columnas"}),
    ("refuerzo_expuesto", {"es": "Refuerzo (varillas) expuesto"}),
    ("pisos_desnivelados", {"es": "Pisos desnivelados o inclinados"}),
    ("separacion_muro_estructura", {"es": "Separación entre muros y estructura"}),
    ("grieta_ancha", {"es": "Grieta llamativamente ancha (fotografiar con escala)"}),
    ("ninguna", {"es": "Ninguna"}),
]

# Indicadores de riesgo no estructural (tabla 3-21) + redes rotas.
# Insumo del revisor y de seguridad para el PMU.
peligros_exterior = [
    ("fachada_balcon", {"es": "Fachada, balcón o antepecho a punto de caer"}),
    ("cubierta_tejas", {"es": "Tejas o elementos de cubierta sueltos"}),
    ("cielo_raso", {"es": "Cielos rasos desprendidos"}),
    ("tanque_elevado", {"es": "Tanque elevado inclinado o suelto"}),
    ("gas", {"es": "Olor o fuga de gas"}),
    ("electricas", {"es": "Cables o postes eléctricos caídos"}),
    ("quimicos", {"es": "Derrame de químicos o tóxicos"}),
    ("escombros_via", {"es": "Escombros sobre la vía"}),
    ("ninguno", {"es": "Ninguno"}),
]

CHOICES = [
    ("confirmaciones", confirmaciones),
    ("tipos_inspeccion", tipos_inspeccion),
    ("motivos_no_inspeccion", motivos_no_inspeccion),
    ("tipo_zona", tipo_zona),
    ("via_tipo", via_tipo),
    ("sistemas_estructurales", sistemas_estructurales),
    ("tipos_entrepiso", tipos_entrepiso),
    ("usos", usos),
    ("rangos_ano", rangos_ano),
    ("colapsos", colapsos),
    ("inclinaciones", inclinaciones),
    ("asentamientos", asentamientos),
    ("fallas_talud", fallas_talud),
    ("morfologias", morfologias),
    ("habitada_opts", habitada_opts),
    ("niveles_dano", niveles_dano),
    ("danos_globales", danos_globales),
    ("senales", senales),
    ("peligros_exterior", peligros_exterior),
]


def survey_columns():
    cols = list(PLAIN_COLS)
    for c in I18N_COLS:
        cols += [f"{c}::{lang}" for lang in LANGS.values()]
    return cols


def survey_row(row):
    out = [row.get(c, "") for c in PLAIN_COLS]
    for c in I18N_COLS:
        val = row.get(c, {})
        out += [val.get(code, "") for code in LANGS]
    return out


wb = Workbook()

ws = wb.active
ws.title = "survey"
ws.append(survey_columns())
for row in survey:
    ws.append(survey_row(row))

ws = wb.create_sheet("choices")
ws.append(["list_name", "name"]
          + [f"label::{lang}" for lang in LANGS.values()])
for list_name, items in CHOICES:
    for name, label in items:
        ws.append([list_name, name]
                  + [label.get(code, "") for code in LANGS])

ws = wb.create_sheet("settings")
ws.append(["form_title", "form_id", "version", "default_language", "style"])
ws.append(["SafeTag — Formulario Único AIS (v2)", "safetag_ais", "2026081702",
           DEFAULT_LANG, "pages"])

out = Path(__file__).parent / "ais.xlsx"
wb.save(out)
print(f"ok → {out}")
